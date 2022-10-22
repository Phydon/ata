import datetime as dt
from openpyxl import load_workbook


def main():
    workbook = load_workbook(filename="C:/Users/Pohl/main/Rust/ata/test.xlsx")
    sheet = workbook.active

    max_row = 0
    for row in sheet.iter_rows(values_only=True):
        max_row += 1

    sheet.insert_rows(idx=max_row + 1)

    start = dt.datetime.now()
    date = start.strftime("%d.%m.%Y")
    time = start.strftime("%H:%M")
    
    sheet.cell(row=max_row + 1, column=1, value=date)
    starttime_cell = sheet.cell(row=max_row + 1, column=2, value=time)
    starttime_loc = starttime_cell.coordinate
    starttime_formula = f"=IF(ISBLANK({starttime_loc}),\"\",{starttime_loc})"
    sheet.cell(row=max_row + 1, column=3, value=starttime_formula)
    
    random_start = dt.datetime(2022, 10, 22, 12, 10, 23, 1)
    sheet.cell(row=max_row + 1, column=4, value=random_start.strftime("%H:%M"))
    first_break = dt.datetime.now()
    delta = round((first_break - random_start).total_seconds() / (60 * 60), 2)
    sheet.cell(row=max_row + 1, column=5, value=delta)
        
    workbook.save(filename="C:/Users/Pohl/main/Rust/ata/test.xlsx")
    
    
if __name__ == "__main__":
    main()
